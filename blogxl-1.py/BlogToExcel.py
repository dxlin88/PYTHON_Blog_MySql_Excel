from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment
import re
import requests
import json
import time
import datetime

wenzhangID=[ ]
wenzhangName=[ ]

print("Reading Blog.....")

#读Blog页面
for i in range(1,5):
    url="http://blog.sina.com.cn/s/article_sort_1313777975_10001_"+str(i)+".html"
   
    html=requests.get(url)
    strr=html.content.decode()
    
    pat2=re.compile(r'id="t_10001_\w{10}(.*?)"')
    pat1=re.compile(r'.html">(.*?)</a></div>')
   
    namelist=pat1.findall(strr)
    idlist=pat2.findall(strr)

    wenzhangName.extend(namelist)
    wenzhangID.extend(idlist)

#做出文章ID表
IDall=','.join(wenzhangID)

#根据文章ID表读出相应的HTML并洗出点击率
url="http://comet.blog.sina.com.cn/api?maintype=num&uid=4e4ea937&aids="+IDall
html=requests.get(url)
strr=html.content.decode()

pat3=re.compile(r'se\W\W\W\W(.*?)\)')
pat3dict=pat3.findall(strr)

patt= pat3dict[0]
data=json.loads(patt)

print("Processing Date to xlsBook" )
print("")

#wb = Workbook()

#打开excell表
wb =load_workbook("blog20190902.xlsx")
ws = wb.active
ws.title = datetime.datetime.now().strftime('%Y-%m')
day = int(datetime.datetime.now().strftime('%d'))
weekday = datetime.datetime.now().strftime('%a,%d')
time = datetime.datetime.now().strftime('%H:%M:%S')

#填表格内容置中设置
al=Alignment(horizontal='center',vertical='center')
bd= Border(left=Side(border_style='thin',color='ff000000'),
           right=Side(border_style='thin',color='FF000000'),
           top=Side(border_style='thin',color='FF000000'),
           bottom=Side(border_style='thin',color='FF000000')
            )

for i in range(1,len(wenzhangName)+6):
    ws.cell(i,day+2).alignment=al
    ws.cell(i,day+2).border=bd
#填填表日期和时间
ws.cell(1,day+2).value = weekday
ws.cell(2,day+2).value = time

#根据清洗后内容填Excell表
TotalCount=0
for i in range(0,len(wenzhangName)):
    ws.cell(i+3,1).value = wenzhangName[i].replace('&nbsp;',' ')
    ws.cell(i+3,day+2).value = data[wenzhangID[i]]["r"]
    TotalCount=ws.cell(i+3,day+2).value+TotalCount
# 比前一天有增加变换底色
    if ws.cell(i+3,day+2).value-ws.cell(i+3,day+1).value > 0:     
        print(wenzhangName[i].replace('&nbsp;',' ')+'  '+ "+"+str(ws.cell(i+3,day+2).value-(ws.cell(i+3,day+1).value))) 
        ws.cell(i+3,day+2).fill = PatternFill(fgColor="FFC7CE", fill_type = "solid")

#总结字段的格式设置和内容填入
ws.cell(len(wenzhangName)+3,1).value = "总阅读点击数"
ws.cell(len(wenzhangName)+3,1).fill = PatternFill(fgColor="F2F2F2", fill_type = "solid")
ws.cell(len(wenzhangName)+4,1).value = "前一天阅读点击数"
ws.cell(len(wenzhangName)+4,1).fill = PatternFill(fgColor="F2F2F2", fill_type = "solid")
ws.cell(len(wenzhangName)+5,1).value = "比前一天阅读点数增减"
ws.cell(len(wenzhangName)+5,1).fill = PatternFill(fgColor="F2F2F2", fill_type = "solid")
ws.cell(len(wenzhangName)+3,day+2).value=TotalCount
ws.cell(len(wenzhangName)+3,day+2).fill = PatternFill(fgColor="F2F2F2", fill_type = "solid")
TotalAdd=ws.cell(len(wenzhangName)+3,day+2).value - int(ws.cell(len(wenzhangName)+3,day+1).value)
ws.cell(len(wenzhangName)+4,day+2).value=TotalAdd
ws.cell(len(wenzhangName)+4,day+2).fill = PatternFill(fgColor="F2F2F2", fill_type = "solid")
LoopRate=int(ws.cell(len(wenzhangName)+4,day+2).value)-int(ws.cell(len(wenzhangName)+4,day+1).value)
ws.cell(len(wenzhangName)+5,day+2).value=LoopRate
ws.cell(len(wenzhangName)+5,day+2).fill = PatternFill(fgColor="F2F2F2", fill_type = "solid")
print("")
print("TotalCount is ",TotalCount)
print("YesterdayAdd is ",TotalAdd)
print("LoopRate is ",LoopRate)
print("")
wb.save("blog20190902.xlsx")
print("End of processe")
input("QUIT Press <enter>")